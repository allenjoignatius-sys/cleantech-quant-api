"""
Enterprise Excel financial-model export (openpyxl).

Generates a fully formatted, *working* workbook — not a static dump. The LCOH,
P&L and balance-sheet sheets are built from live Excel formulas that reference a
single ``Inputs`` sheet, so a banker can open the file, change an assumption and
watch every downstream number (and the balance-sheet tie-out) recompute. This
mirrors the maths in :mod:`app.quant.lcoh` exactly.
"""
from __future__ import annotations

import io
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.quant.lcoh import LCOHInputs, PolicyConfig, calculate_lcoh

# ── House style ────────────────────────────────────────────────────────────────
INK = "0D1117"
ACCENT = "1F6FEB"
LIME = "2EA043"
HEADER_FILL = PatternFill("solid", fgColor=INK)
SUBHEAD_FILL = PatternFill("solid", fgColor="161B22")
INPUT_FILL = PatternFill("solid", fgColor="0B3D2E")
TOTAL_FILL = PatternFill("solid", fgColor="13315C")
WHITE = Font(color="FFFFFF", name="Calibri")
WHITE_BOLD = Font(color="FFFFFF", bold=True, name="Calibri")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=16, name="Calibri")
MONEY = '#,##0.00'
MONEY0 = '#,##0'
PCT = '0.0%'
USD_KG = '$#,##0.000'
THIN = Side(style="thin", color="30363D")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


@dataclass
class ModelAssumptions:
    """P&L / balance-sheet assumptions layered on top of the LCOH inputs."""
    h2_sale_price_usd_per_kg: float = 5.50
    debt_fraction: float = 0.60
    debt_interest_rate: float = 0.06
    tax_rate: float = 0.21


# Stable cell addresses on the Inputs sheet (label col A, value col B, unit col C).
_INPUT_ROWS = {
    "capex_usd_per_kw": 4,
    "electrolyzer_capacity_mw": 5,
    "electricity_price_usd_per_mwh": 6,
    "efficiency_kwh_per_kg": 7,
    "capacity_factor": 8,
    "plant_life_years": 9,
    "discount_rate": 10,
    "fixed_om_pct_capex": 11,
    "stack_replacement_pct_capex": 12,
    "stack_lifetime_hours": 13,
    "water_usd_per_kg": 14,
    "other_opex_usd_per_kg": 15,
    "h2_sale_price_usd_per_kg": 16,
    "debt_fraction": 17,
    "debt_interest_rate": 18,
    "tax_rate": 19,
    "hours_per_year": 20,
}


def _ref(name: str) -> str:
    return f"Inputs!$B${_INPUT_ROWS[name]}"


def _title(ws, text: str, span: int = 6) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=text)
    c.font = TITLE_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26


def _build_inputs(ws, inp: LCOHInputs, asm: ModelAssumptions) -> None:
    _title(ws, "CleanTech Quant — Green H₂ Model · Inputs")
    hdr = ["Parameter", "Value", "Unit"]
    for j, h in enumerate(hdr, start=1):
        c = ws.cell(row=3, column=j, value=h)
        c.font = WHITE_BOLD
        c.fill = SUBHEAD_FILL
    rows = [
        ("capex_usd_per_kw", "Electrolyzer system CAPEX", inp.capex_usd_per_kw, "$/kW", MONEY),
        ("electrolyzer_capacity_mw", "Nameplate capacity", inp.electrolyzer_capacity_mw, "MW", MONEY0),
        ("electricity_price_usd_per_mwh", "Electricity price", inp.electricity_price_usd_per_mwh, "$/MWh", MONEY),
        ("efficiency_kwh_per_kg", "System efficiency", inp.efficiency_kwh_per_kg, "kWh/kg", MONEY),
        ("capacity_factor", "Capacity factor", inp.capacity_factor, "fraction", PCT),
        ("plant_life_years", "Plant life", inp.plant_life_years, "years", MONEY0),
        ("discount_rate", "Discount rate (WACC)", inp.discount_rate, "fraction", PCT),
        ("fixed_om_pct_capex", "Fixed O&M", inp.fixed_om_pct_capex, "% CAPEX/yr", PCT),
        ("stack_replacement_pct_capex", "Stack replacement", inp.stack_replacement_pct_capex, "% CAPEX", PCT),
        ("stack_lifetime_hours", "Stack lifetime", inp.stack_lifetime_hours, "hours", MONEY0),
        ("water_usd_per_kg", "Water cost", inp.water_usd_per_kg, "$/kg", USD_KG),
        ("other_opex_usd_per_kg", "Other OPEX", inp.other_opex_usd_per_kg, "$/kg", USD_KG),
        ("h2_sale_price_usd_per_kg", "H₂ sale price", asm.h2_sale_price_usd_per_kg, "$/kg", USD_KG),
        ("debt_fraction", "Debt fraction", asm.debt_fraction, "fraction", PCT),
        ("debt_interest_rate", "Debt interest rate", asm.debt_interest_rate, "fraction", PCT),
        ("tax_rate", "Corporate tax rate", asm.tax_rate, "fraction", PCT),
        ("hours_per_year", "Hours per year", 8760, "h", MONEY0),
    ]
    for key, label, value, unit, fmt in rows:
        r = _INPUT_ROWS[key]
        ws.cell(row=r, column=1, value=label).font = Font(name="Calibri")
        vc = ws.cell(row=r, column=2, value=value)
        vc.fill = INPUT_FILL
        vc.font = WHITE_BOLD
        vc.number_format = fmt
        vc.border = BORDER
        ws.cell(row=r, column=3, value=unit).font = Font(italic=True, color="8B949E")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14


def _build_lcoh(ws) -> None:
    _title(ws, "Levelised Cost of Hydrogen — live build-up")
    ws.cell(row=3, column=1, value="Component").font = WHITE_BOLD
    ws.cell(row=3, column=1).fill = SUBHEAD_FILL
    ws.cell(row=3, column=2, value="USD/kg H₂").font = WHITE_BOLD
    ws.cell(row=3, column=2).fill = SUBHEAD_FILL

    cap_kw = f"({_ref('electrolyzer_capacity_mw')}*1000)"
    annual_kg = f"({cap_kw}*{_ref('hours_per_year')}*{_ref('capacity_factor')}/{_ref('efficiency_kwh_per_kg')})"
    capex_total = f"({_ref('capex_usd_per_kw')}*{cap_kw})"
    crf = (f"({_ref('discount_rate')}*(1+{_ref('discount_rate')})^{_ref('plant_life_years')}/"
           f"((1+{_ref('discount_rate')})^{_ref('plant_life_years')}-1))")

    # Annual production helper line
    ws.cell(row=4, column=1, value="Annual H₂ production (kg)")
    ws.cell(row=4, column=2, value=f"={annual_kg}").number_format = MONEY0

    # CAPEX/kg
    ws.cell(row=5, column=1, value="Annualised CAPEX")
    ws.cell(row=5, column=2, value=f"=({capex_total}*{crf})/{annual_kg}").number_format = USD_KG
    # Fixed O&M
    ws.cell(row=6, column=1, value="Fixed O&M")
    ws.cell(row=6, column=2, value=f"=({capex_total}*{_ref('fixed_om_pct_capex')})/{annual_kg}").number_format = USD_KG
    # Stack replacement
    op_hours = f"({_ref('hours_per_year')}*{_ref('capacity_factor')}*{_ref('plant_life_years')})"
    n_repl = f"MAX(0,({op_hours}/{_ref('stack_lifetime_hours')})-1)"
    ws.cell(row=7, column=1, value="Stack replacement")
    ws.cell(row=7, column=2,
            value=f"=({capex_total}*{_ref('stack_replacement_pct_capex')}*{n_repl}*{crf})/{annual_kg}"
            ).number_format = USD_KG
    # Electricity
    ws.cell(row=8, column=1, value="Electricity")
    ws.cell(row=8, column=2,
            value=f"={_ref('efficiency_kwh_per_kg')}*({_ref('electricity_price_usd_per_mwh')}/1000)"
            ).number_format = USD_KG
    # Water / Other
    ws.cell(row=9, column=1, value="Water")
    ws.cell(row=9, column=2, value=f"={_ref('water_usd_per_kg')}").number_format = USD_KG
    ws.cell(row=10, column=1, value="Other OPEX")
    ws.cell(row=10, column=2, value=f"={_ref('other_opex_usd_per_kg')}").number_format = USD_KG

    # Total LCOH
    tc = ws.cell(row=11, column=1, value="LCOH (USD/kg H₂)")
    tc.font = WHITE_BOLD
    tc.fill = TOTAL_FILL
    tv = ws.cell(row=11, column=2, value="=SUM(B5:B10)")
    tv.font = WHITE_BOLD
    tv.fill = TOTAL_FILL
    tv.number_format = USD_KG
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16


def _build_pnl(ws, years: int) -> None:
    _title(ws, "Profit & Loss (USD) — annual projection", span=years + 1)
    cap_kw = f"({_ref('electrolyzer_capacity_mw')}*1000)"
    annual_kg = f"({cap_kw}*{_ref('hours_per_year')}*{_ref('capacity_factor')}/{_ref('efficiency_kwh_per_kg')})"
    capex_total = f"({_ref('capex_usd_per_kw')}*{cap_kw})"

    labels = [
        "Revenue", "Electricity cost", "Fixed O&M", "Other cash OPEX",
        "EBITDA", "Depreciation", "EBIT", "Interest", "EBT", "Tax", "Net income",
    ]
    # Year header row
    ws.cell(row=3, column=1, value="Line item / Year").font = WHITE_BOLD
    ws.cell(row=3, column=1).fill = SUBHEAD_FILL
    for y in range(1, years + 1):
        c = ws.cell(row=3, column=1 + y, value=f"Y{y}")
        c.font = WHITE_BOLD
        c.fill = SUBHEAD_FILL
        c.alignment = Alignment(horizontal="right")
    row_of = {name: 4 + i for i, name in enumerate(labels)}
    for name, r in row_of.items():
        ws.cell(row=r, column=1, value=name).font = (
            WHITE_BOLD if name in ("EBITDA", "EBIT", "Net income") else Font(name="Calibri")
        )

    for y in range(1, years + 1):
        col = get_column_letter(1 + y)
        c = lambda name: f"{col}{row_of[name]}"  # noqa: E731
        ws[c("Revenue")] = f"={annual_kg}*{_ref('h2_sale_price_usd_per_kg')}"
        ws[c("Electricity cost")] = f"=-{annual_kg}*{_ref('efficiency_kwh_per_kg')}*({_ref('electricity_price_usd_per_mwh')}/1000)"
        ws[c("Fixed O&M")] = f"=-{capex_total}*{_ref('fixed_om_pct_capex')}"
        ws[c("Other cash OPEX")] = f"=-{annual_kg}*({_ref('water_usd_per_kg')}+{_ref('other_opex_usd_per_kg')})"
        ws[c("EBITDA")] = f"=SUM({col}{row_of['Revenue']}:{col}{row_of['Other cash OPEX']})"
        ws[c("Depreciation")] = f"=-{capex_total}/{_ref('plant_life_years')}"
        ws[c("EBIT")] = f"={c('EBITDA')}+{c('Depreciation')}"
        ws[c("Interest")] = f"=-{capex_total}*{_ref('debt_fraction')}*{_ref('debt_interest_rate')}"
        ws[c("EBT")] = f"={c('EBIT')}+{c('Interest')}"
        ws[c("Tax")] = f"=-MAX(0,{c('EBT')})*{_ref('tax_rate')}"
        ws[c("Net income")] = f"={c('EBT')}+{c('Tax')}"
        for name in labels:
            cell = ws[c(name)]
            cell.number_format = MONEY0
            if name in ("EBITDA", "EBIT", "Net income"):
                cell.font = WHITE_BOLD
                cell.fill = TOTAL_FILL
    ws.column_dimensions["A"].width = 20
    for y in range(1, years + 1):
        ws.column_dimensions[get_column_letter(1 + y)].width = 14


def _build_balance_sheet(ws, years: int) -> None:
    _title(ws, "Balance Sheet (USD) — must tie out to zero", span=years + 1)
    cap_kw = f"({_ref('electrolyzer_capacity_mw')}*1000)"
    capex_total = f"({_ref('capex_usd_per_kw')}*{cap_kw})"
    dep = f"({capex_total}/{_ref('plant_life_years')})"

    labels = [
        "Net PP&E", "Cash", "Total assets",
        "Debt", "Share capital", "Retained earnings", "Total equity",
        "Total liab. + equity", "Balance check",
    ]
    ws.cell(row=3, column=1, value="Item / Year").font = WHITE_BOLD
    ws.cell(row=3, column=1).fill = SUBHEAD_FILL
    for y in range(0, years + 1):
        c = ws.cell(row=3, column=2 + y, value=("Y0" if y == 0 else f"Y{y}"))
        c.font = WHITE_BOLD
        c.fill = SUBHEAD_FILL
        c.alignment = Alignment(horizontal="right")
    row_of = {name: 4 + i for i, name in enumerate(labels)}
    for name, r in row_of.items():
        bold = name in ("Total assets", "Total liab. + equity", "Balance check")
        ws.cell(row=r, column=1, value=name).font = WHITE_BOLD if bold else Font(name="Calibri")

    for y in range(0, years + 1):
        col = get_column_letter(2 + y)
        prev = get_column_letter(1 + y)  # previous year column
        def c(name):  # noqa: E731
            return f"{col}{row_of[name]}"
        if y == 0:
            ws[c("Net PP&E")] = f"={capex_total}"
            ws[c("Cash")] = 0
            ws[c("Debt")] = f"={capex_total}*{_ref('debt_fraction')}"
            ws[c("Share capital")] = f"={capex_total}*(1-{_ref('debt_fraction')})"
            ws[c("Retained earnings")] = 0
        else:
            ni = f"'P&L'!{get_column_letter(1 + y)}{14}"  # Net income row 14 on P&L
            ws[c("Net PP&E")] = f"=MAX(0,{prev}{row_of['Net PP&E']}-{dep})"
            ws[c("Cash")] = f"={prev}{row_of['Cash']}+{ni}+{dep}"
            ws[c("Debt")] = f"={prev}{row_of['Debt']}"
            ws[c("Share capital")] = f"={prev}{row_of['Share capital']}"
            ws[c("Retained earnings")] = f"={prev}{row_of['Retained earnings']}+{ni}"
        ws[c("Total assets")] = f"={c('Net PP&E')}+{c('Cash')}"
        ws[c("Total equity")] = f"={c('Share capital')}+{c('Retained earnings')}"
        ws[c("Total liab. + equity")] = f"={c('Debt')}+{c('Total equity')}"
        ws[c("Balance check")] = f"={c('Total assets')}-{c('Total liab. + equity')}"
        for name in labels:
            cell = ws[c(name)]
            cell.number_format = MONEY0
            if name in ("Total assets", "Total liab. + equity"):
                cell.font = WHITE_BOLD
                cell.fill = TOTAL_FILL
    ws.column_dimensions["A"].width = 22
    for y in range(0, years + 1):
        ws.column_dimensions[get_column_letter(2 + y)].width = 14


def build_financial_model(
    inputs: LCOHInputs,
    assumptions: ModelAssumptions | None = None,
    policy: PolicyConfig | None = None,
    pnl_years: int = 10,
) -> bytes:
    """
    Build the complete workbook and return it as ``.xlsx`` bytes.

    Sheets: Inputs · LCOH · P&L · Balance Sheet · Summary.
    All downstream sheets use live formulas referencing ``Inputs``.
    """
    asm = assumptions or ModelAssumptions()
    pnl_years = max(1, min(pnl_years, inputs.plant_life_years))

    wb = Workbook()
    ws_in = wb.active
    ws_in.title = "Inputs"
    _build_inputs(ws_in, inputs, asm)
    _build_lcoh(wb.create_sheet("LCOH"))
    _build_pnl(wb.create_sheet("P&L"), pnl_years)
    _build_balance_sheet(wb.create_sheet("Balance Sheet"), pnl_years)

    # Summary sheet with the deterministic Python figure for reference.
    res = calculate_lcoh(inputs, policy)
    ws_sum = wb.create_sheet("Summary")
    _title(ws_sum, "Executive Summary")
    ws_sum.cell(row=3, column=1, value="Headline LCOH (model, USD/kg)").font = WHITE_BOLD
    ws_sum.cell(row=3, column=2, value="='LCOH'!B11").number_format = USD_KG
    ws_sum.cell(row=4, column=1, value="LCOH after policy (Python check)")
    ws_sum.cell(row=4, column=2, value=round(res.lcoh_after_policy_usd_per_kg, 4)).number_format = USD_KG
    ws_sum.cell(row=5, column=1, value="Annual production (t H₂)")
    ws_sum.cell(row=5, column=2, value=round(res.annual_h2_tonnes, 1)).number_format = MONEY0
    ws_sum.column_dimensions["A"].width = 34
    ws_sum.column_dimensions["B"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
