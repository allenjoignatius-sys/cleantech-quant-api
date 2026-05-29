"""Unit tests for the openpyxl financial-model export (no DB/network)."""
import io

import pytest
from openpyxl import load_workbook

from app.quant.lcoh import LCOHInputs, PolicyConfig
from app.quant.excel_export import build_financial_model, ModelAssumptions


@pytest.fixture(scope="module")
def workbook_bytes():
    return build_financial_model(
        LCOHInputs(capex_usd_per_kw=1100, electricity_price_usd_per_mwh=40),
        ModelAssumptions(h2_sale_price_usd_per_kg=6.0),
        PolicyConfig(ira_45v_ptc=True),
        pnl_years=8,
    )


class TestExcelExport:
    def test_returns_nonempty_bytes(self, workbook_bytes):
        assert isinstance(workbook_bytes, bytes)
        assert len(workbook_bytes) > 3000  # a real xlsx zip

    def test_expected_sheets(self, workbook_bytes):
        wb = load_workbook(io.BytesIO(workbook_bytes))
        assert wb.sheetnames == ["Inputs", "LCOH", "P&L", "Balance Sheet", "Summary"]

    def test_inputs_values_written(self, workbook_bytes):
        wb = load_workbook(io.BytesIO(workbook_bytes))
        ws = wb["Inputs"]
        assert ws["B4"].value == 1100        # capex $/kW
        assert ws["B6"].value == 40          # electricity price

    def test_lcoh_total_is_live_formula(self, workbook_bytes):
        wb = load_workbook(io.BytesIO(workbook_bytes))
        ws = wb["LCOH"]
        assert isinstance(ws["B11"].value, str) and ws["B11"].value.startswith("=SUM(")
        # electricity line references the Inputs sheet
        assert "Inputs!" in ws["B8"].value

    def test_pnl_has_formulas_and_years(self, workbook_bytes):
        wb = load_workbook(io.BytesIO(workbook_bytes))
        ws = wb["P&L"]
        assert ws["B3"].value == "Y1"
        assert ws["I3"].value == "Y8"   # 8 year columns
        # Net income row contains a formula
        assert isinstance(ws["B14"].value, str) and ws["B14"].value.startswith("=")

    def test_balance_sheet_check_is_formula(self, workbook_bytes):
        wb = load_workbook(io.BytesIO(workbook_bytes))
        ws = wb["Balance Sheet"]
        # "Balance check" is the last labelled row; it must be a tie-out formula
        labels = [ws.cell(row=r, column=1).value for r in range(4, 13)]
        assert "Balance check" in labels
        check_row = 4 + labels.index("Balance check")
        assert isinstance(ws.cell(row=check_row, column=2).value, str)
        assert ws.cell(row=check_row, column=2).value.startswith("=")

    def test_respects_pnl_year_cap(self):
        # pnl_years cannot exceed plant life
        b = build_financial_model(LCOHInputs(plant_life_years=5), pnl_years=30)
        wb = load_workbook(io.BytesIO(b))
        assert wb["P&L"]["F3"].value == "Y5"
